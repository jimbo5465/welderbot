# ══════════════════════════════════════════════════════════════════════════════
# engine/ncr_excel.py
# تولید فایل Excel رسمی NCR از روی رکورد ncrs در دیتابیس + قالب
# media/templates/NCR.xlsx
#
# قانون معماری (مطابق engine/report_builder.py):
#   - فقط از db.models می‌خواند؛ هیچ وابستگی به context مکالمه تلگرام ندارد.
#   - می‌توان با فقط یک ncr_id و یک شماره، در هر زمانی خروجی ساخت.
#   - مختصات سلول‌ها و منطق تیک‌زدن گزینه‌ها از ncr_sample.py گرفته شده
#     (همان خروجی نمونه تأییدشده در media/exports/NCR_sample_v2.xlsx).
#
# تصمیم فاز ۲: عکس‌های NCR فعلاً وارد اکسل نمی‌شوند (فقط در DB و پوشهٔ
# media/ncr_photos نگهداری می‌شوند). اگر قالب جای عکس داشته باشد در فاز
# بعدی با الگوی _PHOTO_MAX_* در report_builder.py اضافه می‌شود.
# ══════════════════════════════════════════════════════════════════════════════

import os

from openpyxl import load_workbook
from openpyxl.styles import Font

import config
from db.models import get_ncr_by_id, get_project_by_id, get_contractor_by_id

TEMPLATE_PATH = os.path.join(config._PROJECT_ROOT, "media", "templates", "NCR.xlsx")

# فونت ساده و خوانا برای همهٔ مقادیر (مطابق ncr_sample.py)
_SIMPLE_FONT = "Tahoma"
_SIMPLE_SIZE = 11

# گزینه‌های تیک‌داری که در قالب هستند و کاربر از میان آن‌ها انتخاب می‌کند
# حتماً باید با لیست‌های دکمه در handlers/ncr.py یکسان باشند.
OPERATION_TYPES = ["ساختمانی", "نصب", "پیش راه اندازی", "راه اندازی", "سایر"]
DISCIPLINES = ["ساختمان", "مکانیک", "برق", "ابزار دقیق", "سایر"]
CAUSE_TYPES = ["مهندسی", "ساخت و تامین", "نصب", "راه اندازی", "حمل و جابجایی", "انبارش و نگهداری", "سایر"]


def _set_value(ws, coord: str, value: str | None, bold: bool = True, size: int = _SIMPLE_SIZE) -> None:
    cell = ws[coord]
    cell.value = value or ""
    cell.font = Font(name=_SIMPLE_FONT, size=size, bold=bold, color="FF000000")


def _clean_options(ws, coord: str, selected_opt: str | None) -> None:
    """
    مربع‌های وینگدینگ (کاراکتر '5') را حذف و جلوی گزینهٔ انتخابی تیک تمیز
    می‌گذارد — دقیقاً همان منطق ncr_sample.clean_options.
    """
    cell = ws[coord]
    text = cell.value or ""
    text = str(text).replace("5", "")
    if selected_opt:
        idx = text.find(selected_opt)
        if idx != -1:
            text = text[:idx] + "✓ " + selected_opt + text[idx + len(selected_opt):]
    cell.value = text.replace("  ", " ").strip()
    cell.font = Font(name=_SIMPLE_FONT, size=10, bold=False, color="FF000000")


def build_ncr_excel(ncr_id: int, ncr_number: str | None = None) -> str:
    """
    فایل Excel رسمی NCR را برای ضبط مشخص می‌سازد.

    ورودی:
        ncr_id:      شناسهٔ رکورد در جدول ncrs
        ncr_number:  شمارهٔ NCR برای چاپ در فرم — در صورت None از
                     خود رکورد دیتابیس خوانده می‌شود (باید قبلاً submit شود)

    خروجی:
        مسیر مطلق فایل Excel تولید‌شده (در config.NCR_EXPORT_PATH)

    خطا:
        ValueError اگر NCR یا پروژه/پیمانکار یافت نشود
        FileNotFoundError اگر template اصلی موجود نباشد
    """
    ncr = get_ncr_by_id(ncr_id)
    if ncr is None:
        raise ValueError(f"NCR با id={ncr_id} یافت نشد.")

    project = get_project_by_id(ncr["project_id"])
    contractor = get_contractor_by_id(ncr["contractor_id"])
    if project is None:
        raise ValueError(f"پروژهٔ NCR {ncr_id} یافت نشد.")
    if contractor is None:
        raise ValueError(f"پیمانکار NCR {ncr_id} یافت نشد.")

    if not os.path.isfile(TEMPLATE_PATH):
        raise FileNotFoundError(f"تمپلیت یافت نشد: {TEMPLATE_PATH}")

    number = ncr_number or ncr.get("ncr_number") or ""

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ─── هدر ──────────────────────────────────────────────────────────────
    _set_value(ws, "C2", f"پروژه: {project['name']}")
    _set_value(ws, "P2", number)
    _set_value(ws, "P4", ncr.get("reported_date") or "")

    # ─── مشخصات ───────────────────────────────────────────────────────────
    _set_value(ws, "A6", f"نام پیمانکار : {contractor['name']}")
    _set_value(ws, "A7", ncr.get("island") or "")
    _set_value(ws, "F7", ncr.get("unit") or "")
    _set_value(ws, "A8", ncr.get("drawing_number") or "")

    # ─── گزینه‌ها: حذف مربع‌ها + تیک تمیز + فونت ساده ────────────────────
    _clean_options(ws, "H5", ncr.get("operation_type"))       # عملیات
    _clean_options(ws, "J7", ncr.get("discipline"))           # دیسیپلین
    _clean_options(ws, "A11", ncr.get("cause"))               # علت وقوع

    # ─── متن‌ها ───────────────────────────────────────────────────────────
    _set_value(ws, "A9", ncr.get("description") or "")
    _set_value(ws, "A12", ncr.get("corrective_action") or "")
    _set_value(ws, "A15", ncr.get("equipment_description") or "")

    reporter = ncr.get("reporter_name") or ""
    if ncr.get("reporter_title"):
        reporter = f"{reporter} — {ncr['reporter_title']}"
    _set_value(ws, "A16", reporter)

    # ─── ذخیره خروجی ─────────────────────────────────────────────────────
    os.makedirs(config.NCR_EXPORT_PATH, exist_ok=True)
    serial = number.split("-")[-1] if number else str(ncr_id)
    out_filename = f"NCR_{ncr_id:05d}_{serial}.xlsx"
    out_path = os.path.join(config.NCR_EXPORT_PATH, out_filename)
    wb.save(out_path)

    return out_path