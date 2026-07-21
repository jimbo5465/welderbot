# ══════════════════════════════════════════════════════════════════════════════
# engine/report_builder.py
# تولید فایل Excel رسمی WPQ از روی رکورد قطعی صلاحیت (qualification) در دیتابیس.
#
# قانون معماری: این ماژول فقط از db.models می‌خواند (welder_id, project_id و غیره
# را از خود qualification می‌گیرد)؛ هیچ وابستگی به context مکالمه تلگرام ندارد.
# یعنی می‌توان با فقط یک qualification_id، در هر زمانی (حتی ماه‌ها بعد) گزارش
# ساخت — نه فقط بلافاصله بعد از ثبت تست.
# ══════════════════════════════════════════════════════════════════════════════

import os
import shutil
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

# مقداری که به سایز فونت پیش‌فرض سلول (خوانده‌شده از خود template) اضافه می‌شود
_VALUE_FONT_SIZE_BOOST = 2

import config
from db.models import get_qualification_by_id, get_welder_by_id

# ابعاد باکس عکس (P4:R7) بر حسب پیکسل — از روی column width/row height خود
# template محاسبه شده (۳ ستون عرض ۱۳ | ۴ ردیف ارتفاع ۱۹.۵pt)، با کمی حاشیه
# اطمینان تا عکس هرگز از باکس بیرون نزند.
_PHOTO_MAX_WIDTH_PX = 260
_PHOTO_MAX_HEIGHT_PX = 95

# ستون‌هایی که مقدار در آن‌ها نوشته می‌شود و باید wrap+ارتفاع پویا داشته باشند
_WRAP_ROWS = list(range(9, 23))
_WRAP_COLS = ("G", "M")

# مسیر Template اصلاح‌شده (نسخه split‌شده هدر/امضا) — باید یک‌بار در این مسیر قرار گیرد
TEMPLATE_PATH = os.path.join(config._PROJECT_ROOT, "media", "templates", "WPQ_template.xlsx")

# نگاشت نتیجه دیتابیس -> برچسب نمایشی در فرم (Accept / Reject / N.A)
_RESULT_LABELS = {
    "ACC": "Accept",
    "REJ": "Reject",
}

# ردیف‌های جدول نتایج تست: (نام فیلد در extra_data, شماره ردیف)
_TEST_RESULT_ROWS = {
    "visual": 25,   # از visual_groove_result یا visual_fillet_result (هرکدام None نبود)
    "rt": 26,
    "fracture": 27,
    "macro": 28,
    # ردیف ۲۹ (Guided-Bend) عمداً خالی می‌ماند — طبق تصمیم پروژه، داده‌ای برایش وجود ندارد
}

# ستون‌های Accept/Reject/N.A برای هر ردیف نتیجه تست (anchor سلول‌های merge شده)
_RESULT_COLS = {"ACC": "G", "REJ": "K", "NA": "O"}


def _fmt_list(val) -> str:
    """list یا مقدار ساده را برای نمایش در سلول اکسل به رشته تبدیل می‌کند."""
    if val is None:
        return ""
    if isinstance(val, (list, tuple)):
        return "، ".join(str(v) for v in val)
    return str(val)


def _fmt_dual_process(qual: dict, container_key_gtaw: str, container_key_smaw: str, field: str) -> str:
    """
    ترکیب مقدار یک فیلد از دو پردازش GTAW/SMAW (اگر هر دو فعال باشند).
    مثال: current -> "GTAW:DC / SMAW:AC"
    """
    extra = qual.get("extra_data") or {}
    parts = []
    gtaw = extra.get(container_key_gtaw)
    smaw = extra.get(container_key_smaw)
    if gtaw and gtaw.get(field):
        parts.append(f"GTAW:{gtaw[field]}")
    if smaw and smaw.get(field):
        parts.append(f"SMAW:{smaw[field]}")
    return " / ".join(parts) if parts else ""


def _resolve_photo_path(photo_path: str | None) -> str | None:
    """
    photo_path ذخیره‌شده در دیتابیس ممکن است نسبت به MEDIA_PATH یا نسبت به
    project root ذخیره شده باشد (بسته به نسخه کد). هر دو حالت را چک می‌کند.
    """
    if not photo_path:
        return None
    candidates = [
        os.path.join(config.MEDIA_PATH, os.path.basename(photo_path)),
        os.path.join(config._PROJECT_ROOT, photo_path),
        photo_path,
    ]
    for c in candidates:
        if os.path.isfile(c):
            return c
    return None


def _mark_test_result(ws, row: int, result_value: str | None) -> None:
    """
    روی سلول Accept/Reject/N.A متناظر با نتیجه، علامت ✓ اضافه می‌کند.
    result_value: 'ACC' / 'REJ' / None (یعنی N.A یا هنوز مشخص نیست)
    """
    label = _RESULT_LABELS.get(result_value, "N.A")
    key = result_value if result_value in _RESULT_COLS else "NA"
    col = _RESULT_COLS[key]
    cell = ws[f"{col}{row}"]
    cell.value = f"✓ {label}"
    _bold_value(ws, f"{col}{row}")


def _bold_value(ws, coord: str, extra_size: int = _VALUE_FONT_SIZE_BOOST) -> None:
    """
    فونت یک سلولِ مقدار (نه لیبل) را بولد و کمی بزرگ‌تر می‌کند، بدون تغییر
    فونت پایه‌ی template (رنگ/نام فونت template حفظ می‌شود).
    """
    cell = ws[coord]
    base = cell.font
    cell.font = Font(
        name=base.name,
        size=(base.size or 10) + extra_size,
        bold=True,
        color=base.color,
    )


def build_wpq_excel(qualification_id: int) -> str:
    """
    فایل Excel رسمی WPQ را برای یک صلاحیت مشخص می‌سازد.

    ورودی:
        qualification_id: شناسه رکورد در جدول qualifications

    خروجی:
        مسیر مطلق فایل Excel تولید‌شده (در config.EXCEL_EXPORT_PATH)

    خطا:
        ValueError اگر qualification یا welder یافت نشود
        FileNotFoundError اگر template اصلی موجود نباشد
    """
    qual = get_qualification_by_id(qualification_id)
    if qual is None:
        raise ValueError(f"صلاحیتی با id={qualification_id} یافت نشد.")

    welder = get_welder_by_id(qual["welder_id"])
    if welder is None:
        raise ValueError(f"جوشکار با id={qual['welder_id']} یافت نشد.")

    if not os.path.isfile(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"Template یافت نشد: {TEMPLATE_PATH}\n"
            "باید نسخه اصلاح‌شده (split‌شده) فرم WPQ یک‌بار در این مسیر قرار گیرد."
        )

    extra = qual.get("extra_data") or {}

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["WPQ"]

    # ─── هدر (ردیف ۴ تا ۷) ──────────────────────────────────────────────
    ws["E4"] = f"WQT-{qualification_id:05d}"
    ws["M4"] = extra.get("contractor_name", "")
    ws["E5"] = welder.get("full_name", "")
    ws["M5"] = qual.get("test_date", "")
    ws["E6"] = extra.get("welder_id_no", "")
    ws["M6"] = extra.get("wqt_no", "")
    ws["E7"] = extra.get("coupon_no", "")
    ws["M7"] = extra.get("wps_no", "")
    for coord in ("E4", "M4", "E5", "M5", "E6", "M6", "E7", "M7"):
        _bold_value(ws, coord)

    # ─── عکس جوشکار (P4:R7) ─────────────────────────────────────────────
    # نکته مهم: عکس‌های واقعی (موبایل) معمولاً چند هزار پیکسل هستند.
    # باید قبل از insert، نسبت مقیاس را طوری محاسبه کنیم که عکس داخل
    # باکس بماند (contain-fit) و هرگز از مرز آن بیرون نزند.
    photo_path = _resolve_photo_path(welder.get("photo_path"))
    if photo_path:
        img = XLImage(photo_path)
        orig_w, orig_h = img.width, img.height
        scale = min(_PHOTO_MAX_WIDTH_PX / orig_w, _PHOTO_MAX_HEIGHT_PX / orig_h)
        img.width = round(orig_w * scale)
        img.height = round(orig_h * scale)
        ws.add_image(img, "P4")

    # ─── جدول متغیرهای QW-350 (ردیف ۹ تا ۲۲) ───────────────────────────
    ws["G9"] = qual.get("process", "");                  ws["M9"] = _fmt_list(qual.get("qr_process"))
    ws["G10"] = qual.get("backing", "");                  ws["M10"] = _fmt_list(qual.get("qr_backing"))
    ws["G11"] = qual.get("base_metal_p_no", "");           ws["M11"] = _fmt_list(qual.get("qr_p_no"))
    diam_thick = " / ".join(filter(None, [
        f"OD:{qual['pipe_od_mm']}mm" if qual.get("pipe_od_mm") else "",
    ]))
    ws["G12"] = diam_thick;                                ws["M12"] = f"{_fmt_list(qual.get('qr_diameter'))} | {_fmt_list(qual.get('qr_thickness'))}"
    ws["G13"] = qual.get("filler_f_no", "");                ws["M13"] = _fmt_list(qual.get("qr_f_no"))
    ws["G14"] = _fmt_dual_process(qual, "filler_gtaw", "filler_smaw", "sfa")
    ws["G15"] = qual.get("filler_aws_class", "")
    ws["G16"] = qual.get("deposit_groove_mm", "")
    # M16: رنج تأیید ضخامت رسوب جوش = ۲× ضخامت واقعی نمونه (طبق QW-451).
    # ⚠️ اگر سقف بالایی خاصی هم طبق استاندارد پروژه اعمال می‌شود، اینجا باید اضافه شود.
    if qual.get("deposit_groove_mm"):
        ws["M16"] = round(qual["deposit_groove_mm"] * 2, 2)
    ws["G17"] = qual.get("deposit_fillet_mm", "")
    ws["G18"] = qual.get("test_position", "");              ws["M18"] = f"{_fmt_list(qual.get('qr_position_groove'))} {_fmt_list(qual.get('qr_position_fillet'))}".strip()
    ws["G19"] = extra.get("progression", "")
    ws["M19"] = extra.get("qr_progression", "")
    ws["G20"] = extra.get("shielding_gas", "")
    ws["G21"] = _fmt_dual_process(qual, "elec_gtaw", "elec_smaw", "current")
    ws["G22"] = _fmt_dual_process(qual, "elec_gtaw", "elec_smaw", "polarity")
    for coord in (
        "G9", "M9", "G10", "M10", "G11", "M11", "G12", "M12", "G13", "M13",
        "G14", "G15", "G16", "M16", "G17", "G18", "M18", "G19", "M19", "G20", "G21", "G22",
    ):
        _bold_value(ws, coord)

    # ─── wrap + ارتفاع پویا برای جدول متغیرها (جلوگیری از overflow متن بلند) ───
    for row in _WRAP_ROWS:
        max_len = 0
        for col in _WRAP_COLS:
            cell = ws[f"{col}{row}"]
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        # تخمین تعداد خط لازم (~35 کاراکتر در هر خط با عرض ستون فعلی) و تنظیم ارتفاع
        lines_needed = max(1, -(-max_len // 35))  # ceil division
        if lines_needed > 1:
            ws.row_dimensions[row].height = max(
                ws.row_dimensions[row].height or 19.5, lines_needed * 14
            )

    # ─── نتایج تست (ردیف ۲۵ تا ۲۸؛ ردیف ۲۹ همیشه خالی) ──────────────────
    visual_result = extra.get("visual_groove_result") or extra.get("visual_fillet_result")
    _mark_test_result(ws, 25, visual_result)
    _mark_test_result(ws, 26, extra.get("rt_result"))
    _mark_test_result(ws, 27, extra.get("fracture_result"))
    _mark_test_result(ws, 28, extra.get("macro_result"))

    # ─── ردیف ۳۱: تست‌شده توسط ───────────────────────────────────────────
    signer_name = qual.get("signer_name") or ""
    signer_title = qual.get("signer_title") or ""
    ws["O31"] = f"{signer_name} — {signer_title}".strip(" —")
    _bold_value(ws, "O31")

    # ─── امضای MAPNA-MD1 (NAME + DATE) ──────────────────────────────────
    # نام امضاکننده با فونت بزرگ‌تر (هم‌سطح تیتر) و بولد نمایش داده می‌شود
    ws["I35"] = signer_name
    _bold_value(ws, "I35", extra_size=4)
    ws["I37"] = qual.get("test_date", "")

    # ─── ذخیره خروجی ─────────────────────────────────────────────────────
    os.makedirs(config.EXCEL_EXPORT_PATH, exist_ok=True)
    out_filename = f"WPQ_{welder['national_id']}_{qualification_id}.xlsx"
    out_path = os.path.join(config.EXCEL_EXPORT_PATH, out_filename)
    wb.save(out_path)

    return out_path
