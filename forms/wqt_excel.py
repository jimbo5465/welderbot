"""
ماژول wqt_excel — تولید فایل Excel گواهینامه WQT.
از openpyxl==3.1.2 استفاده می‌کند.
پیاده‌سازی فاز ۷ — امضا از CONTRACTS.md قفل شده است.
"""

from __future__ import annotations

import json
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

import config
from db.models import get_qualification_by_id, get_welder_by_id, list_projects
from utils.dates import gregorian_to_jalali


def generate_wqt(qualification_id: int) -> str:
    """
    فایل Excel گواهینامه WQT را برای یک صلاحیت مشخص تولید می‌کند.

    ورودی:
        qualification_id: شناسه ردیف در جدول qualifications

    خروجی:
        مسیر کامل فایل Excel تولیدشده
        فرمت نام: 'media/exports/WQT_{welder_national_id}_{qualification_id}.xlsx'

    خطا:
        FileNotFoundError اگر qualification_id وجود نداشته باشد
        ValueError اگر داده‌های لازم ناقص باشند
    """
    # ── دریافت داده‌ها از DB ─────────────────────────────────────────────────
    qual = get_qualification_by_id(qualification_id)
    if not qual:
        raise FileNotFoundError(
            f"صلاحیت با شناسه {qualification_id} یافت نشد."
        )

    welder = get_welder_by_id(qual["welder_id"])
    if not welder:
        raise ValueError(f"جوشکار مرتبط با صلاحیت {qualification_id} یافت نشد.")

    # دریافت نام پروژه
    projects = list_projects(active_only=False)
    project  = next((p for p in projects if p["id"] == qual["project_id"]), None)
    project_name = project["name"] if project else "—"

    # ── تبدیل تاریخ‌ها به جلالی ─────────────────────────────────────────────
    def _j(g: str | None) -> str:
        if not g:
            return "—"
        try:
            return gregorian_to_jalali(g)
        except Exception:
            return g or "—"

    test_date_j   = _j(qual.get("test_date"))
    expiry_date_j = _j(qual.get("expiry_date"))

    # ── آماده‌سازی لیست‌های qr_* ─────────────────────────────────────────────
    def _lst(val) -> str:
        if isinstance(val, list):
            return "، ".join(val)
        if isinstance(val, str):
            try:
                return "، ".join(json.loads(val))
            except Exception:
                return val
        return "—"

    # ── ساخت فایل Excel ─────────────────────────────────────────────────────
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "WQT"

    # تنظیم راست‌چین برای زبان فارسی
    ws.sheet_view.rightToLeft = True

    # ── استایل‌ها ────────────────────────────────────────────────────────────
    _bold       = Font(bold=True, size=11)
    _header     = Font(bold=True, size=13, color="FFFFFF")
    _center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _right      = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    _fill_blue  = PatternFill("solid", fgColor="1F4E79")
    _fill_gray  = PatternFill("solid", fgColor="D9E1F2")
    _thin       = Side(style="thin")
    _border     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def _cell(row: int, col: int, value, bold=False, fill=None, align=None):
        """نوشتن مقدار در سلول با استایل."""
        c = ws.cell(row=row, column=col, value=value)
        c.border = _border
        if bold:
            c.font = _bold
        if fill:
            c.fill  = fill
            c.font  = _header
        c.alignment = align or _right
        return c

    # ── سطر عنوان اصلی ───────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = "گواهینامه صلاحیت جوشکار — ASME Section IX"
    title_cell.font      = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill      = _fill_blue
    title_cell.alignment = _center
    title_cell.border    = _border
    ws.row_dimensions[1].height = 30

    # ── بخش اول: اطلاعات جوشکار و پروژه ─────────────────────────────────────
    row = 3
    headers_info = [
        ("نام جوشکار",    welder.get("full_name", "—")),
        ("کد ملی",        welder.get("national_id", "—")),
        ("پروژه",         project_name),
        ("تاریخ آزمون",   test_date_j),
        ("تاریخ انقضا",   expiry_date_j),
        ("امضاکننده",     f"{qual.get('signer_name', '—')} — {qual.get('signer_title', '—')}"),
    ]
    ws.merge_cells(f"A{row}:F{row}")
    _cell(row, 1, "اطلاعات شناسایی", bold=True, fill=_fill_blue, align=_center)
    ws.row_dimensions[row].height = 22
    row += 1

    for label, value in headers_info:
        _cell(row, 1, label, bold=True, fill=_fill_gray)
        ws.merge_cells(f"B{row}:F{row}")
        _cell(row, 2, value)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # فاصله

    # ── بخش دوم: متغیرهای آزمون ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    _cell(row, 1, "متغیرهای آزمون", bold=True, fill=_fill_blue, align=_center)
    ws.row_dimensions[row].height = 22
    row += 1

    test_vars = [
        ("فرآیند جوشکاری",  qual.get("process", "—")),
        ("Backing",           qual.get("backing", "—")),
        ("نوع اتصال",        qual.get("joint_type", "—")),
        ("نوع نمونه",        qual.get("specimen_type", "—")),
        ("قطر لوله (mm)",    str(qual.get("pipe_od_mm") or "—")),
        ("P-Number فلز پایه",qual.get("base_metal_p_no", "—")),
        ("F-Number الکترود", qual.get("filler_f_no", "—")),
        ("کلاس AWS",         qual.get("filler_aws_class") or "—"),
        ("ضخامت رسوب Groove (mm)", str(qual.get("deposit_groove_mm") or "—")),
        ("ضخامت رسوب Fillet (mm)", str(qual.get("deposit_fillet_mm") or "—")),
        ("تعداد پاس",        str(qual.get("pass_count", "—"))),
        ("موقعیت آزمون",     qual.get("test_position", "—")),
    ]
    for label, value in test_vars:
        _cell(row, 1, label, bold=True, fill=_fill_gray)
        ws.merge_cells(f"B{row}:F{row}")
        _cell(row, 2, value)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # ── بخش سوم: دامنه صلاحیت (۸ فیلد qr_*) ──────────────────────────────────
    ws.merge_cells(f"A{row}:F{row}")
    _cell(row, 1, "دامنه صلاحیت محاسبه‌شده — ASME Sec. IX", bold=True, fill=_fill_blue, align=_center)
    ws.row_dimensions[row].height = 22
    row += 1

    qr_fields = [
        ("فرآیند واجد شرایط",            qual.get("qr_process", "—")),
        ("Backing واجد شرایط",            qual.get("qr_backing", "—")),
        ("P-Number واجد شرایط",           _lst(qual.get("qr_p_no"))),
        ("محدوده ضخامت",                  qual.get("qr_thickness", "—")),
        ("محدوده قطر",                    qual.get("qr_diameter", "—")),
        ("موقعیت‌های Groove واجد شرایط",  _lst(qual.get("qr_position_groove"))),
        ("موقعیت‌های Fillet واجد شرایط",  _lst(qual.get("qr_position_fillet"))),
        ("F-Number واجد شرایط",           _lst(qual.get("qr_f_no"))),
    ]
    for label, value in qr_fields:
        _cell(row, 1, label, bold=True, fill=_fill_gray)
        ws.merge_cells(f"B{row}:F{row}")
        _cell(row, 2, value)
        ws.row_dimensions[row].height = 20
        row += 1

    # ── تنظیم عرض ستون‌ها ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # ── ذخیره فایل ───────────────────────────────────────────────────────────
    os.makedirs(config.EXCEL_EXPORT_PATH, exist_ok=True)
    national_id = welder.get("national_id", "unknown")
    filename    = f"WQT_{national_id}_{qualification_id}.xlsx"
    output_path = os.path.join(config.EXCEL_EXPORT_PATH, filename)
    wb.save(output_path)

    return output_path
