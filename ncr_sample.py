import os

from openpyxl import load_workbook
from openpyxl.styles import Font

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(PROJECT_ROOT, "media", "templates", "NCR.xlsx")
OUT_PATH = os.path.join(PROJECT_ROOT, "media", "exports", "NCR_sample_v2.xlsx")

# فونت ساده و خوانا برای همهٔ مقادیر
SIMPL_FONT = "Tahoma"
SIMPL_SIZE = 11


def set_value(ws, coord, value, bold=True, size=SIMPL_SIZE):
    cell = ws[coord]
    cell.value = value
    cell.font = Font(name=SIMPL_FONT, size=size, bold=bold, color="FF000000")


def clean_options(ws, coord, selected_opt):
    """مربع‌های وینگدینگز (کاراکتر 5) را حذف و تیک تمیز جلوی گزینهٔ انتخابی می‌گذارد.
    فونت‌های عجیب تکه‌تکه را با یک فونت ساده جایگزین می‌کند."""
    cell = ws[coord]
    text = cell.value or ""
    # حذف تمام کاراکترهای تیکِ وینگدینگ (در این تمپلیت به صورت '5' ذخیره‌شده‌اند)
    text = text.replace("5", "")
    # افزودن ✓ جلوی گزینهٔ انتخاب‌شده
    idx = text.find(selected_opt)
    if idx != -1:
        text = text[:idx] + "✓ " + selected_opt + text[idx + len(selected_opt):]
    cell.value = text.replace("  ", " ").strip()
    cell.font = Font(name=SIMPL_FONT, size=10, bold=False, color="FF000000")


def build_sample():
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ─── هدر ──────────────────────────────────────────────────────────────
    set_value(ws, "C2", "پروژه: نیروگاه سیکل ترکیبی شیراز")
    set_value(ws, "P2", "SHAZ-NCR-1404-001")
    set_value(ws, "P4", "1404/05/12")

    # ─── مشخصات ─────────────────────────────────────────────────────────────
    set_value(ws, "A6", "نام پیمانکار : شرکت پیمانکار نیروگاهی پارس")
    set_value(ws, "A7", "جزیره یک (توربین گاز)")
    set_value(ws, "F7", "واحد شماره یک")
    set_value(ws, "A8", "D-001 / Rev: B")

    # ─── گزینه‌ها: حذف مربع‌ها + تیک تمیز + فونت ساده ─────────────────────
    clean_options(ws, "H5", "نصب")       # عملیات
    clean_options(ws, "J7", "مکانیک")    # دیسیپلین
    clean_options(ws, "A11", "نصب")      # علت وقوع

    # ─── متن‌ها ─────────────────────────────────────────────────────────────
    set_value(ws, "A9", "عدم انطباق در فیت‌آپ و اتصال فلنج‌های خط ۲ اینچ به منیفولد اصلی — فاصله محوری خارج از تلورانس استاندارد پروژه می‌باشد.")
    set_value(ws, "A12", "بازگشت سطح فلنج‌ها به وضعیت استاندارد و انجام بازرسی مجدد همراه با انجام VT")
    set_value(ws, "A15", "تورک رنچ برقی، آچارها، اینسترومنت کالیبره")
    set_value(ws, "A16", "مهندس محمد رضایی — بازرس مکانیک")

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    if os.path.exists(OUT_PATH):
        os.remove(OUT_PATH)
    wb.save(OUT_PATH)
    print("saved:", OUT_PATH)


if __name__ == "__main__":
    build_sample()